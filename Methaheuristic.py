import random
import math
from typing import List, Union
from openpyxl import workbook
from pathlib import Path
from numpy import cos, pi, exp, sqrt, prod, sin
import numpy as np
import string
import openpyxl as op
import timeit
import argparse

import os
from src import problem
from src.problem.SUKP import SUKP


class Metaheuristic:

    numberIndividual = 0
    numberIteration = 0
    numberField = 1
    consultationFactor = 0.0
    numberOfVariables = 0
    numberOfVariablesToChange = 0
    iteracionMejorFitness = 0
    k1 = 0.0
    k2 = 0.0
    poblation = []
    fields = []
    numAcceptedMoves = 0
    numRejectedMoves = 0
    rpd = 0.0

    def __init__(self, numberindividual: int, numberiteration: int,
                 type: string, worksheet, problem: SUKP):
        self.numberIndividual = numberindividual
        self.numberIteration = numberiteration
        self.type = type
        self.problem: SUKP = problem
        self.ws = worksheet

    def run(self):
        self.ws.cell(row=1,
                     column=1,
                     value="Resultados metaheuristica con instancia " +
                     self.type)
        self.generateInitialPoblation()
        self.ordenarhienas()
        h = 5
        e = self.generaRandom()
        for i in range(len(e)):
            e[i] = 2 * e[i] * h - h
        b = self.generaRandom()
        for i in range(len(b)):
            b[i] = 2 * b[i]

        iteration = 0
        ch = []
        mejorRana = self.poblation[0]
        mejorFitness = self.opt(self.poblation[0])
        self.ws.cell(row=3, column=1, value="Iteracion")
        self.ws.cell(row=3, column=2, value="Divers")
        self.ws.cell(row=3, column=3, value="inten")
        self.ws.cell(row=3, column=4, value="noMejora")
        for i in range(len(self.poblation)):
            self.ws.cell(row=3, column=5 + i, value="fitness indiv" + str(i))
        self.ws.cell(row=4, column=1, value="Inicial")
        for i in range(len(self.poblation)):
            self.ws.cell(row=4,
                         column=i + 5,
                         value=self.opt(self.poblation[i]))

        while iteration < self.numberIteration:
            inten = 0
            diver = 0
            no_mejor = 0
            print(str(iteration) + " -> " + str(mejorFitness))

            for i in range(len(self.poblation)):
                pk = [0 for k in range(len(self.poblation[i]))]
                indiv_inten = random.randint(1, 5)
                inten += indiv_inten
                for _ in range(indiv_inten):
                    for j in range(len(mejorRana)):
                        pk[j] = self.discretiza(
                            self.transforma(mejorRana[j] -
                                            (e[j] * ((b[j] * mejorRana[j]) -
                                                     self.poblation[i][j]))))
                pk = self.problem.fixSolution(pk)
                if self.opt(pk) > self.opt(self.poblation[i]):
                    ch.append(pk)
                else:
                    pk = self.poblation[i]
                    ch.append(pk)

            for i in range(len(ch)):
                actualrana = [0 for k in range(len(ch[i]))]
                indiv_diver = random.randint(1, 5)
                diver += indiv_diver
                for j in range(len(mejorRana)):
                    for _ in range(indiv_diver):
                        actualrana[j] = self.discretiza(
                            self.transforma(ch[i][j] / len(ch)))
                mejorRana = self.problem.fixSolution(mejorRana)
                if self.opt(actualrana) > self.opt(ch[i]):
                    self.poblation[i] = actualrana
                else:
                    if self.opt(ch[i]) == self.opt(self.poblation[i]):
                        no_mejor += 1
                    self.poblation[i] = ch[i]
            h = 5 - iteration / self.numberIteration
            e = self.generaRandom()
            for i in range(len(e)):
                e[i] = 2 * e[i] * h - h
            b = self.generaRandom()
            for i in range(len(b)):
                b[i] = 2 * b[i]
            self.ordenarhienas()
            if self.opt(self.poblation[0]) > mejorFitness:
                mejorRana = self.poblation[0]
                mejorFitness = self.opt(self.poblation[0])
                self.iteracionMejorFitness = iteration
            iteration += 1
            ch = []

            self.ws.cell(row=4 + iteration, column=1, value=iteration)
            self.ws.cell(row=4 + iteration, column=2, value=diver / 100)
            self.ws.cell(row=4 + iteration, column=3, value=inten / 100)
            self.ws.cell(row=4 + iteration, column=4, value=no_mejor)
            for i in range(len(self.poblation)):
                self.ws.cell(row=4 + iteration,
                             column=i + 5,
                             value=self.opt(self.poblation[i]))

    def opt(self, x):
        return self.problem.fitness(x)

    def ordenarhienas(self):
        orden = False
        i = 0
        while (i < len(self.poblation)) and (orden == False):
            i += 1
            orden = True
            for j in range(len(self.poblation) - 1):
                if self.opt(self.poblation[j + 1]) > self.opt(
                        self.poblation[j]):
                    orden = False
                    aux = self.poblation[j]
                    self.poblation[j] = self.poblation[j + 1]
                    self.poblation[j + 1] = aux
        return

    def generaRandom(self) -> List[int]:
        randomSolution = [
            random.random() for i in range(len(self.problem.values))
        ]
        return randomSolution

    def discretiza(self, x):
        t = random.random()
        if t < x:
            return 1
        else:
            return 0

    def transforma(self, var):
        t = math.tan(var)
        if t < 0:
            t *= -1
        return t

    def generateInitialPoblation(self):
        self.poblation = []
        for _ in range(self.numberIndividual):
            self.poblation.append(self.problem.generateInitialSolution())


numberPoblation = 100
numberIterations = 5000

initial_path = "C:\\Users\\Acer\\np_meta\\SUKP"
os.chdir(initial_path)
set = os.listdir()

for i in range(len(set)):
    path = os.path.join(initial_path, set[i])
    os.chdir(path)
    instances = os.listdir()
    for j in range(len(instances)):
        newpath = os.path.join(initial_path, set[i], instances[j])
        base, ext = os.path.splitext(instances[j])
        if (ext != ".txt"):
            continue
        problem = SUKP()
        print("problema " + instances[j])
        problem.read_file(newpath)
        wb = op.Workbook()
        ws1 = wb.active
        ws1.title = "iteracion 1"
        ws2 = wb.create_sheet("Iteracion 2")
        ws3 = wb.create_sheet("Iteracion 3")
        ws4 = wb.create_sheet("Iteracion 4")
        ws5 = wb.create_sheet("Iteracion 5")
        ws6 = wb.create_sheet("Iteracion 6")
        ws7 = wb.create_sheet("Iteracion 7")
        ws8 = wb.create_sheet("Iteracion 8")
        ws9 = wb.create_sheet("Iteracion 9")
        ws10 = wb.create_sheet("Iteracion 10")
        ws11 = wb.create_sheet("Iteracion 11")
        ws12 = wb.create_sheet("Iteracion 12")
        ws13 = wb.create_sheet("Iteracion 13")
        ws14 = wb.create_sheet("Iteracion 14")
        ws15 = wb.create_sheet("Iteracion 15")
        ws16 = wb.create_sheet("Iteracion 16")
        ws17 = wb.create_sheet("Iteracion 17")
        ws18 = wb.create_sheet("Iteracion 18")
        ws19 = wb.create_sheet("Iteracion 19")
        ws20 = wb.create_sheet("Iteracion 20")
        ws21 = wb.create_sheet("Iteracion 21")
        ws22 = wb.create_sheet("Iteracion 22")
        ws23 = wb.create_sheet("Iteracion 23")
        ws24 = wb.create_sheet("Iteracion 24")
        ws25 = wb.create_sheet("Iteracion 25")
        ws26 = wb.create_sheet("Iteracion 26")
        ws27 = wb.create_sheet("Iteracion 27")
        ws28 = wb.create_sheet("Iteracion 28")
        ws29 = wb.create_sheet("Iteracion 29")
        ws30 = wb.create_sheet("Iteracion 30")
        ws31 = wb.create_sheet("Iteracion 31")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws1, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        valtime = (start - end)
        print(valtime)
        ws1.cell(row=2, column=2, value=valtime)
        wb.save("Results" + base + ".xlsx")

        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws2, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws2.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws3, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws3.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws4, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws4.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws5, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws5.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws6, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws6.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws7, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws7.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws8, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws8.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws9, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws9.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws10, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws10.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws11, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws11.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws12, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws12.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws13, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws13.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws14, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws14.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws15, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws15.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws16, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws16.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws17, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws17.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws18, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws18.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws19, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws19.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws20, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws20.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws21, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws21.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws22, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws22.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws23, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws23.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws24, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws24.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws25, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws25.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws26, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws26.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws27, problem)
        mstart = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws27.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws28, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws28.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws29, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws29.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws30, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws30.cell(row=2, column=2, value=(start - end))
        wb.save("Results" + base + ".xlsx")
        metaheuristic = Metaheuristic(numberPoblation, numberIterations,
                                      instances[j], ws31, problem)
        start = timeit.timeit()
        metaheuristic.run()
        end = timeit.timeit()
        ws31.cell(row=2, column=2, value=(start - end))

        wb.save("Results" + base + ".xlsx")
